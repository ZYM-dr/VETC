import matplotlib.pyplot as plt
import numpy as np
import math
from openpyxl import load_workbook
# 图片热力图生成
##读取路径
wb = load_workbook(filename=r'heatmapdata.xlsx')
##读取名字为Sheet1的sheet表
ws = wb.get_sheet_by_name("Sheet2")
arr1 = []
arr2 = []
## 遍历第2行到10000行
for row_A in range(2, 171):
    ## 遍历第2行到10000行，第1列
    a1 = ws.cell(row=row_A, column=1).value
    ## 遍历第2行到10000行，第3列
    a2 = ws.cell(row=row_A, column=2).value
    if a1:
        # 写入数组1
        #arr1.append(a1)
        arr1.append(-a2)
    if a2:
        # 写入数组2
        #arr2.append(a2)
        arr2.append(-a1)

# POINT DATASET

x = arr1
y = arr2

# x = [20, 28, 15, 20, 18, 25, 15, 18, 18, 20, 25, 30, 25, 22, 30, 22, 38, 40, 38, 30, 22, 20, 35, 33, 35]
# y = [20, 14, 15, 20, 15, 20, 32, 33, 45, 50, 20, 20, 20, 25, 30, 38, 20, 28, 33, 50, 48, 40, 30, 35, 36]
# # DEFINE GRID SIZE AND RADIUS(h)
grid_size = 1
h = 10

# GETTING X,Y MIN AND MAX
x_min = min(x)
x_max = max(x)
y_min = min(y)
y_max = max(y)

# CONSTRUCT GRID
x_grid = np.arange(x_min - h, x_max + h, grid_size)
y_grid = np.arange(y_min - h, y_max + h, grid_size)
x_mesh, y_mesh = np.meshgrid(x_grid, y_grid)

# GRID CENTER POINT
xc = x_mesh + (grid_size / 2)
yc = y_mesh + (grid_size / 2)


# FUNCTION TO CALCULATE INTENSITY WITH QUARTIC KERNEL
def kde_quartic(d, h):
    dn = d / h
    P = (15 / 16) * (1 - dn ** 2) ** 2
    return P


# PROCESSING
intensity_list = []
for j in range(len(xc)):
    intensity_row = []
    for k in range(len(xc[0])):
        kde_value_list = []
        for i in range(len(x)):
            # CALCULATE DISTANCE
            d = math.sqrt((xc[j][k] - x[i]) ** 2 + (yc[j][k] - y[i]) ** 2)
            if d <= h:
                p = kde_quartic(d, h)
            else:
                p = 0
            kde_value_list.append(p)
        # SUM ALL INTENSITY VALUE
        p_total = sum(kde_value_list)
        intensity_row.append(p_total)
    intensity_list.append(intensity_row)

# HEATMAP OUTPUT
intensity = np.array(intensity_list)
plt.pcolormesh(x_mesh, y_mesh, intensity)
# plt.plot(x, y, 'ro')
plt.colorbar()
plt.show()